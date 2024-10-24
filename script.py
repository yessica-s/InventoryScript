import sys
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from product import Product
from enum import Enum

############################################################
#                     GLOBAL VARIABLES                     #
############################################################

# Track what type of data user wants to format
class TimePeriod(Enum):
    MONTH = "M"
    ANNUAL = "A"

# Initialise variables to hold indices for particular columns
sku_index = None
quarter_index = None
product_index = None
invoice_date_index = None
customer_index = None
quantity_index = None
sale_index = None 
profit_index = None
document_index = None

# Start of headings e.g. SKU, Product is LibreOffice row 6, Script row 4
heading_row_index = 4

# Initialise list of products
products = []

# Time period user is enterring data for 
time_period = None

def check_path(file_path: str):
    """Check if the file path is valid."""

    if not os.path.exists(file_path):
        print(f"Invalid file path: {file_path}")
        sys.exit(1)

def load_data(file_path: str):
    """Load data from XLSX file."""
    try:
        df = pd.read_excel(file_path)
        print("Data loaded successfully.")
        return df
    except Exception as e:
        print(f"Error loading data: {e}")
        return None
    
"""
Note:
- Heading column does not have index
- Row index 0 starts at row "2" in LibreOffice
- Assuming standard formatting as in "annualreport.xlsx" file, 
  the correct headings row should begin at Row 4 in Script
"""

def set_column_indices(data):

    global sku_index, quarter_index, product_index, invoice_date_index
    global customer_index, quantity_index, sale_index, profit_index, document_index

    # Get column indices for required column headings
    row_values = data.iloc[heading_row_index] # Retrieve row as list

    # Iterate through column headings and store respective indices
    for index, value in enumerate(row_values): 
        match value:
            case "SKU":
                sku_index = index
            case "Quarter":
                quarter_index = index
            case "Product":
                product_index = index
            case "Invoice date":
                invoice_date_index = index
            case "Customer":
                customer_index = index
            case "Quantity":
                quantity_index = index
            case "Sale":
                sale_index = index
            case "Profit":
                profit_index = index
            case "Document #":
                document_index = index
            case _: # something else
                pass     
    return

def find_all_totals(data) -> dict:
    """Aggregates Sales, Quantity, Profit per Product Code (SKU)"""
    global products
    product_totals = {}

    for index, row in data.iterrows(): # Loop through rows
        if index < heading_row_index + 1: # skip row that do not contain sales data
            continue

        # get all the transaction info for the row
        current_sku = row.iloc[sku_index]
        current_description = row.iloc[product_index] if product_index is not None else None
        new_quantity = row.iloc[quantity_index] if quantity_index is not None else 0
        new_sales = row.iloc[sale_index] if sale_index is not None else 0
        new_profit = row.iloc[profit_index] if profit_index is not None else 0
        new_invoice_date = row.iloc[invoice_date_index] if invoice_date_index is not None else ""
        new_document = row.iloc[document_index] if document_index is not None else ""

        if current_sku not in product_totals:
            product_totals[current_sku] = Product(current_sku)
            product_totals[current_sku].set_description(current_description)
            products.append(product_totals[current_sku]) # append product object to list of products

        # increment the current product totals with the new values
        date = None
        if time_period == TimePeriod.MONTH:
            date = "Month"
        else:
            date = "Year"
            
        product_totals[current_sku].update_product(new_quantity, new_sales, new_profit, new_invoice_date, new_document, date)

    return product_totals

def format_product_totals(product_totals: dict):
    """Formats the aggregated product data into a new excel sheet. Handled totals only"""

    filename = "Inventory_Totals.xlsx"

    # Create filepath to save in downloads directory
    save_directory = "/home/medigroup/Documents/Inventory Management"
    file_path = os.path.join(save_directory, filename)

    # If already exists, relabel
    base_name, extension = os.path.splitext(filename)
    counter = 1

    while os.path.exists(file_path):
        file_path = os.path.join(save_directory, f"{base_name}_{counter}{extension}")
        counter += 1

    # Create workbook and add new sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Formatted Data"

    sheet.append(['SKU', 'Total Quantity Sold', 'Total Sales', 'Total Profit'])
    for cell in sheet[1]: # for cell in first row, set to bold
        cell.font = Font(bold = True, name = 'Arial')

    for sku, product in product_totals.items(): # Loop through all products
        # print(product)
        quantity = product.get_total_quantity()
        sales = product.get_total_sales()
        profit = product.get_total_profit()

        row = [sku, quantity, sales, profit]
        
        sheet.append(row) # Add the row to the sheet

    # Set column widths
    for column in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = max_length + 2  # Adding some padding
    workbook.save(file_path)
    print(f"Excel file created: {file_path}")

    return

def get_last_filled_column(sheet, row_number):
    """Return the index of the last filled column in the specified row."""
    last_col = 2
    for col in range(2, sheet.max_column + 1):  # Columns are 1-indexed in openpyxl
        if sheet.cell(row=row_number, column=col).value is not None:
            last_col = col
    return last_col

def get_last_filled_row(sheet, column_number):
    """Return the index of the last filled row in the specified column."""
    last_row = 0
    for row in range(1, sheet.max_row + 1):  # Rows are 1-indexed in openpyxl
        if sheet.cell(row=row, column=column_number).value is not None:
            last_row = row
    return last_row

def find_product_index_in_list(product_sku: str, existing_products: list) -> int | None:
    for i in range(0, len(existing_products)):
        if existing_products[i] == product_sku:
            return i
        
    return None

def insert_products_and_descr(sheet, heading_row: int):
    # for initialising the product codes and names upon creation of the spreadsheet
    row = heading_row + 1
    for product in products: 
        sheet.cell(row, 1, value=product.get_sku())
        sheet.cell(row, 2, value=product.get_description())
        row += 1

def insert_product_data(sheet, product: Product, existing_products: list, quantity_col: int, sale_col: int, profit_col: int, heading_row: int):
    for date, index in product._transaction_years.items():
        product_sku = product.get_sku()
        product_description = product.get_description()
        total_quantity = product._quantity_lists[index]
        total_sales = product._sales_lists[index]
        total_profit = product._profit_lists[index]

        # find row index of product in product column
        row_to_insert_in = find_product_index_in_list(product_sku, existing_products)
        if row_to_insert_in is None: # product doesn't already exist in spreadshet so insert
            row_to_insert_in = get_last_filled_row(sheet, 1) + 1 # get next empty column
            sheet.cell(row_to_insert_in, 1, value=product_sku) # insert sku
            sheet.cell(row_to_insert_in, 2, value=product_description) # insert description
        else:
            row_to_insert_in += 4 # to account for non-heading rows at top

        sheet.cell((heading_row-1), quantity_col, value=date) # add heading of date

        sheet.cell(heading_row, quantity_col, value="Quantity") # add field headings
        sheet.cell(heading_row, sale_col, value="Sales")
        sheet.cell(heading_row, profit_col, value="Profit")

        sheet.cell(row_to_insert_in, quantity_col, value=total_quantity)
        sheet.cell(row_to_insert_in, sale_col, value=total_sales)
        sheet.cell(row_to_insert_in, profit_col, value=total_profit)

        quantity_col += 3 # increment columns for next 
        sale_col += 3
        profit_col += 3


"""
Note: apparently openpyxl columns are 1-indexed not 0-indexed
"""

def add_date_to_sheet():

    file_path = '/home/medigroup/Documents/Inventory Management/Sales.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    sheet = None
    if time_period == TimePeriod.MONTH:
        sheet = workbook['Monthly']
    else: 
        sheet = workbook['Annual']
    existing_products = [cell[0] for cell in sheet.iter_rows(min_col=1, max_col=1, values_only=True)]
    existing_products = existing_products[3:] # discard headings etc. - should be 4???

    heading_row = 3
    max_col = get_last_filled_column(sheet, heading_row)
 
    quantity_col = max_col + 1
    sale_col = max_col + 2
    profit_col = max_col + 3

    # make more robust later i.e. deal with if year found that's alr in spreadsheet
    for product in products: # Loop through products
        insert_product_data(sheet, product, existing_products, quantity_col, sale_col, profit_col, heading_row)

    # insert_products_and_descr(sheet, heading_row)

    workbook.save(file_path)
    print(f"Workbook saved: {file_path}")
    return

def main():
    print("Starting script...")

    if len(sys.argv) != 2: 
        print("Usage error: python3 script.py <path_to_xlsx_file_data>")
        print("Make sure you are in the InventoryScripts Directory when running this command")
        sys.exit(1)

    global time_period
    time_period = input("Please type M for Monthly and A for Annual depending on the spreadsheet you would like to edit\n")
    time_period = time_period.upper()
    time_period = TimePeriod(time_period) # either MONTHLY or ANNUALLY

    file_path = sys.argv[1]  # Path to XLSX Data
    check_path(file_path) # Check if the file path is valid

    # Load the data
    data = load_data(file_path)
    if data is None:
        print("No data found in provided file")
        sys.exit(1)
        
    set_column_indices(data) # Find and assign indices for required columns to global variables
    product_totals = find_all_totals(data)

    # format_product_totals(product_totals) # don't really need
    add_date_to_sheet()

    print("Script Complete")
    sys.exit(1)

if __name__ == "__main__":
    main()
